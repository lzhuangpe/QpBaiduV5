import asyncio
from re import split,search
from random import choice
import datetime
import time
import aiohttp
from pyquery import PyQuery as pq
import tldextract
import openpyxl

def text2list(file):
    with open(file, encoding='utf-8') as f:
        return (line.strip() for line in f.readlines())


def get_search_url_list(keywords, depth=3):
    start_url = 'https://www.baidu.com/s?wd={keyword}&pn={page}&oq={keyword}&ie=utf-8'
    return (start_url.format(keyword=keyword, page=str(page*10)) for keyword in keywords for page in range(0, depth))


async def get_response(url, text=False):
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7',
        'user-agent': choice(user_agents)
    }
    async with aiohttp.ClientSession() as session:
        try:
            async with Semaphore:
                async with session.get(url, headers=headers, timeout=20, verify_ssl=False) as resp:
                    resp.raise_for_status()
                    if not text:
                        return resp
                    else:
                        return await resp.text('utf-8', 'ignore')
        except aiohttp.client_exceptions.ClientResponseError as e:
            print('aiohttp.client_exceptions.ClientResponseError:{erro} {url}'.format(erro=e, url=url))
        except aiohttp.client_exceptions.ClientConnectorError as e:
            print('aiohttp.client_exceptions.ClientConnectorError:{erro} {url}'.format(erro=e, url=url))
        except aiohttp.client_exceptions.ClientOSError as e:
            print('aiohttp.client_exceptions.ClientOSError:{erro} {url}'.format(erro=e, url=url))
        except asyncio.TimeoutError as e:
            print('asyncio.TimeoutError:{erro} {url}'.format(erro=e, url=url))
        except aiohttp.client_exceptions.ServerDisconnectedError as e:
            print('aiohttp.client_exceptions.ServerDisconnectedError:{erro} {url}'.format(erro=e, url=url))
        except aiohttp.client_exceptions.InvalidURL:
            pass


async def get_one_page_index_list(html):
    if not html:
        print('get_one_page_index Error:'.format('resp is not find'))
    doc = pq(html)
    return doc('.result').items()


async def get_c_show_url(html):
    url = html('.c-showurl').text()
    if url:
        o = tldextract.extract(url)
        domain = '{}.{}'.format(o.domain, o.suffix)
        return domain


async def get_sort_url(url):
    if search('https?://', url):
        url = split('/', url)[2]
    else:
        url = split('/', url)[0]
    return url

def save_results(results):
    if results:
        file = '{}/{}.{}'.format('result', str(datetime.date.today()), 'xlsx')
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'search_data'
        # 创建表头
        table_hard = list(results[0].keys())
        for i in range(len(table_hard)):
            sheet.cell(row=1, column=i+1, value=str(table_hard[i]))
        # 写入数据
        for i in range(len(results)):
            for j in range(len(results[i])):
                sheet.cell(row=i+2, column=j+1, value=str(list(results[i].values())[j]))
        # 保存数据
        wb.save(file)


async def main(search_url):
    html = await get_response(url=search_url, text=True)
    if not html:
        return
    items = await get_one_page_index_list(html)
    global counter
    for item in items:
        counter += 1
        show_url = await get_c_show_url(html=item)
        if not show_url or show_url in exclude_domain:
            continue
        href = str(item('h3 a').attr('href'))
        title = item('h3 a').text()
        resp = await get_response(href)
        if not resp:
            continue
        url = str(resp.url)
        sort_url = await get_sort_url(url)
        if sort_url in exclude_url or sort_url in cache_url:
            continue
        cache_url.append(sort_url)
        # html = pq(await get_response(url, text=True))
        # title = html('title').text()
        result = {'title': title,
                  'show_url': show_url,
                  'url': url,
                  'href': '{}{}'.format(href.strip(), '&wd=&eqid='),
                  'search_url': search_url
                  }
        print(counter, result)
        results.append(result)

if __name__ == '__main__':
    counter = 0
    results, cache_url = list(), list()
    Semaphore = asyncio.Semaphore(100)
    urls = get_search_url_list(keywords=text2list('./text/keyword.txt'))
    exclude_domain = set(text2list('./text/exclude_domain.txt'))
    exclude_url = set(text2list('./text/exclude_url.txt'))
    user_agents = list(text2list('./text/user_agent.txt'))
    tasks = [main(url) for url in urls]
    try:
        start = time.time()
        loop = asyncio.get_event_loop()
        loop.run_until_complete(asyncio.wait(tasks))
        end = time.time()
        print('总共用时:{}'.format(str(end - start)))
    finally:
        loop.close()
        save_results(results)